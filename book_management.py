import openpyxl
from datetime import date, timedelta, datetime
file_name = 'book_details.xlsx'


class Book:
    def __init__(self, id, name, author, publisher):
        self.id = id
        self.name = name
        self.author = author
        self.publisher = publisher

    def issue(self):
        today = date.today()
        issue_date = today
        return_date = today + timedelta(days=7)
        self.issue_date = issue_date.strftime("%Y-%m-%d")
        self.return_date = return_date.strftime("%Y-%m-%d")


def get_books():
    global file_name
    wb = openpyxl.load_workbook(file_name, data_only=True)
    sheet = wb['Sheet1']
    max_row = sheet.max_row + 1
    books = []

    for row in range(3, max_row):
        data = []

        for col in range(1,5):
            data.append(sheet.cell(row=row, column=col).value)

        r_date = sheet.cell(row=row, column=6).value
        try:
            r_date = datetime.strptime(str(r_date),"%Y-%m-%d")
        except:
            r_date = datetime.strptime(str(r_date),"%Y-%m-%d %H:%M:%S")
        if r_date.date() <= date.today():
            books.append(Book(*data))

    return books


def issue_book(book):
    global file_name
    wb = openpyxl.load_workbook(file_name, data_only=True)
    sheet = wb['Sheet1']
    max_row = sheet.max_row + 1
    book.issue()

    for row in range(3, max_row):
        if book.id == sheet.cell(row=row, column=1).value:
            sheet.cell(row=row,column=5).value = book.issue_date
            sheet.cell(row=row,column=6).value = book.return_date
            wb.save(file_name)
            msg = '\n\tBook issued successfully!'
            msg += f'\nIssue date : {book.issue_date}'
            msg += f'\nReturn date : {book.return_date}'
            return msg
            
    else:
        return 'Book not found!'
